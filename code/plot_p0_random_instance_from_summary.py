from pathlib import Path
import argparse
from datetime import datetime
import random
import subprocess
import sys

from openpyxl import load_workbook


TOOL_DIR = Path(__file__).resolve().parent
DEFAULT_SUMMARY_XLSX = TOOL_DIR / "p0_192_instance_parameter_summary.xlsx"
GRID_SCRIPT = TOOL_DIR / "plot_p0_all_methods_runs_grid.py"
OVERLAY_SCRIPT = TOOL_DIR / "plot_p0_all_methods_pf_overlay.py"
HEATMAP_SCRIPT = TOOL_DIR / "plot_p0_all_methods_pf_heatmap.py"
EAF_SCRIPT = TOOL_DIR / "plot_p0_all_methods_eaf.py"
DEFAULT_OUTPUT_ROOT = Path(r"C:\Users\yiting\Desktop\NCHU\lab\TEVC\code\output")


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Randomly select instance(s) from p0_192_instance_parameter_summary.xlsx, "
            "then generate all-methods PF plots."
        )
    )
    parser.add_argument(
        "--summary-xlsx",
        type=Path,
        default=DEFAULT_SUMMARY_XLSX,
        help="Excel summary file with an Instances sheet.",
    )
    parser.add_argument(
        "--split",
        choices=["train", "validation", "test"],
        default=None,
        help="Optional split filter before random sampling.",
    )
    parser.add_argument(
        "--instance-name",
        default=None,
        help="Use a specific instance name instead of random sampling. Overrides --num-instances.",
    )
    parser.add_argument(
        "--num-instances",
        type=int,
        default=1,
        help="Number of instances to randomly sample without replacement.",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Random seed for reproducible sampling.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help="Root output folder. Overlay PNGs go to <root>/overlay; heatmap PNGs go to <root>/heatmap.",
    )
    parser.add_argument(
        "--batch-id",
        default=None,
        help="Optional output batch folder name. Defaults to a timestamp so previous outputs are not overwritten.",
    )
    parser.add_argument(
        "--run-count",
        type=int,
        default=30,
        help="Number of runs passed to both plotting scripts.",
    )
    parser.add_argument(
        "--x-obj",
        type=int,
        default=1,
        help="1-based objective index for x-axis.",
    )
    parser.add_argument(
        "--y-obj",
        type=int,
        default=2,
        help="1-based objective index for y-axis.",
    )
    parser.add_argument(
        "--raw",
        action="store_true",
        help="Pass --raw to both plotting scripts.",
    )
    parser.add_argument(
        "--method-normalize-grid",
        action="store_true",
        help="Pass --method-normalize to the 2x3 grid script only.",
    )
    parser.add_argument(
        "--no-lines-overlay",
        action="store_true",
        help="Pass --no-lines to the single overlay script only.",
    )
    parser.add_argument(
        "--heatmap-bins",
        type=int,
        default=60,
        help="Number of bins per axis passed to the heatmap script.",
    )
    parser.add_argument(
        "--linear-heatmap",
        action="store_true",
        help="Use raw counts in the heatmap. This is the default; kept for compatibility.",
    )
    parser.add_argument(
        "--log-heatmap",
        action="store_true",
        help="Use log1p(count) in the heatmap instead of raw counts.",
    )
    parser.add_argument(
        "--skip-heatmap",
        action="store_true",
        help="Only generate the two overlay figures and skip the heatmap.",
    )
    parser.add_argument(
        "--eaf-grid-size",
        type=int,
        default=120,
        help="Number of grid points per axis passed to the EAF script.",
    )
    parser.add_argument(
        "--skip-eaf",
        action="store_true",
        help="Skip the EAF band figure.",
    )
    return parser.parse_args()


def normalize_windows_path(path_text: str) -> Path:
    text = str(path_text).strip().strip('"')
    if len(text) >= 4 and text[0] in {"/", "\\"} and text[2] == ":":
        text = text[1:]
    return Path(text)


def read_instances(summary_xlsx: Path):
    if not summary_xlsx.exists():
        raise FileNotFoundError(f"Summary workbook not found: {summary_xlsx}")

    wb = load_workbook(summary_xlsx, read_only=True, data_only=True)
    if "Instances" not in wb.sheetnames:
        raise ValueError(f"Workbook has no 'Instances' sheet: {summary_xlsx}")

    ws = wb["Instances"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    required = {"Split", "Instance name", "K folder", "Instance path"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Instances sheet is missing columns: {sorted(missing)}")

    instances = []
    for row in rows:
        item = dict(zip(headers, row))
        if not item.get("Instance name"):
            continue
        instances.append(item)
    return instances


def choose_instances(instances, split=None, instance_name=None, seed=None, num_instances=1):
    candidates = instances

    if split:
        candidates = [row for row in candidates if row.get("Split") == split]

    if instance_name:
        matches = [row for row in candidates if row.get("Instance name") == instance_name]
        if not matches:
            raise ValueError(f"No instance matched: {instance_name}")
        return [matches[0]]

    if not candidates:
        raise ValueError("No candidate instances after applying filters.")
    if num_instances < 1:
        raise ValueError("--num-instances must be >= 1.")
    if num_instances > len(candidates):
        raise ValueError(
            f"--num-instances={num_instances} exceeds candidate count {len(candidates)}."
        )

    rng = random.Random(seed)
    return rng.sample(candidates, num_instances)


def resolve_k_instance_dir(instance_row) -> Path:
    instance_path = normalize_windows_path(instance_row["Instance path"])
    k_folder = str(instance_row.get("K folder") or "").strip()

    if k_folder:
        k_dir = instance_path / k_folder
        if k_dir.exists():
            return k_dir

    k_dirs = sorted(p for p in instance_path.glob("K_*") if p.is_dir())
    if not k_dirs:
        raise FileNotFoundError(f"No K_* folder found under: {instance_path}")
    return k_dirs[0]


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 10_000):
        candidate = path.with_name(f"{stem}_{index:03d}{suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Could not find a unique filename for {path}")


def run_plot_script(
    script: Path,
    instance_dir: Path,
    args,
    output_dir: Path,
    output_prefix: str,
    extra_flags=None,
):
    if not script.exists():
        raise FileNotFoundError(f"Plot script not found: {script}")

    command = [
        sys.executable,
        str(script),
        "--instance-dir",
        str(instance_dir),
        "--run-count",
        str(args.run_count),
        "--x-obj",
        str(args.x_obj),
        "--y-obj",
        str(args.y_obj),
    ]

    command.extend(["--output-dir", str(output_dir)])
    command.append("--png-only")
    if args.raw:
        command.append("--raw")
    if extra_flags:
        command.extend(extra_flags)

    try:
        result = subprocess.run(
            command,
            check=True,
            text=True,
            capture_output=True,
        )
    except subprocess.CalledProcessError as exc:
        print("\nPlot script failed:")
        print(" ".join(command))
        if exc.stdout:
            print("\nstdout:")
            print(exc.stdout)
        if exc.stderr:
            print("\nstderr:")
            print(exc.stderr)
        raise
    outputs = []
    for line in result.stdout.splitlines():
        text = line.strip()
        if not text:
            continue
        output = Path(text)
        if output.exists() and output_prefix:
            renamed = unique_path(output.with_name(f"{output_prefix}{output.name}"))
            output.rename(renamed)
            outputs.append(str(renamed))
        else:
            outputs.append(text)
    return outputs


def default_batch_id() -> str:
    return datetime.now().strftime("batch_%Y%m%d_%H%M%S_%f")


def safe_folder_name(name: str) -> str:
    keep = []
    for ch in str(name):
        if ch.isalnum() or ch in {"-", "_", "."}:
            keep.append(ch)
        else:
            keep.append("_")
    cleaned = "".join(keep).strip("._")
    return cleaned or "instance"


def run_for_instance(selected, args, batch_id: str):
    instance_dir = resolve_k_instance_dir(selected)
    instance_folder = safe_folder_name(selected["Instance name"])
    output_prefix = f"{batch_id}_{instance_folder}_"
    overlay_dir = args.output_root / "overlay"
    heatmap_dir = args.output_root / "heatmap"
    eaf_dir = args.output_root / "eaf"

    print(f"\nSelected split: {selected['Split']}")
    print(f"Selected instance: {selected['Instance name']}")
    print(f"Selected K folder: {instance_dir}")
    print(f"Output file prefix: {output_prefix}")
    print(f"Overlay output folder: {overlay_dir}")
    print(f"Heatmap output folder: {heatmap_dir}")
    print(f"EAF output folder: {eaf_dir}")

    grid_flags = []
    if args.method_normalize_grid:
        grid_flags.append("--method-normalize")

    overlay_flags = []
    if args.no_lines_overlay:
        overlay_flags.append("--no-lines")

    print("\nRunning 2x3 all-methods runs grid...")
    grid_outputs = run_plot_script(
        GRID_SCRIPT, instance_dir, args, overlay_dir, output_prefix, grid_flags
    )
    for output in grid_outputs:
        print(output)

    print("\nRunning all-methods PF overlay...")
    overlay_outputs = run_plot_script(
        OVERLAY_SCRIPT, instance_dir, args, overlay_dir, output_prefix, overlay_flags
    )
    for output in overlay_outputs:
        print(output)

    if not args.skip_heatmap:
        heatmap_flags = ["--bins", str(args.heatmap_bins)]
        if args.linear_heatmap:
            heatmap_flags.append("--linear")
        elif args.log_heatmap:
            heatmap_flags.append("--log-counts")

        print("\nRunning all-methods PF heatmap...")
        heatmap_outputs = run_plot_script(
            HEATMAP_SCRIPT, instance_dir, args, heatmap_dir, output_prefix, heatmap_flags
        )
        for output in heatmap_outputs:
            print(output)

    if not args.skip_eaf:
        eaf_flags = ["--grid-size", str(args.eaf_grid_size)]

        print("\nRunning all-methods EAF band...")
        eaf_outputs = run_plot_script(
            EAF_SCRIPT, instance_dir, args, eaf_dir, output_prefix, eaf_flags
        )
        for output in eaf_outputs:
            print(output)


def main():
    args = parse_args()

    instances = read_instances(args.summary_xlsx)
    selected_instances = choose_instances(
        instances,
        split=args.split,
        instance_name=args.instance_name,
        seed=args.seed,
        num_instances=args.num_instances,
    )
    batch_id = safe_folder_name(args.batch_id) if args.batch_id else default_batch_id()

    print(f"Output root: {args.output_root}")
    print(f"Batch id: {batch_id}")
    print(f"Instance count: {len(selected_instances)}")

    for index, selected in enumerate(selected_instances, start=1):
        print(f"\n=== Instance {index}/{len(selected_instances)} ===")
        run_for_instance(selected, args, batch_id)


if __name__ == "__main__":
    main()
